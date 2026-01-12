import csv
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from model.metadata_container import MetadataContainer


class MetadataFileIO:
    """Handle reading and writing metadata values from/to Excel files."""

    def read_metadata_values(
            self, metadata_container: MetadataContainer, file_path: Path
    ) -> MetadataContainer | None:
        """Populate metadata `values fields using the given Excel file."""
        workbook: Workbook = load_workbook(file_path,
                                           read_only=True,
                                           data_only=True)
        missing_codes = list(set(metadata_container.codes()))

        try:
            try:
                sheet = workbook[metadata_container.sheet_name]
            except KeyError:
                print(
                    f"Sheet '{metadata_container.sheet_name}' not found in"
                    f" '{file_path}'. Available sheets:"
                    f" {', '.join(workbook.sheetnames)}"
                )
                return metadata_container

            target_column = metadata_container.column_index
            if target_column < 1:
                print(
                    f"column_index must be a positive integer;"
                    f" got {target_column}"
                )
                return metadata_container

            for code in missing_codes:
                metadata = metadata_container.get_metadata(code)
                if metadata is None:
                    continue

                try:
                    cell = sheet[code]
                except ValueError:
                    print(
                        f"Invalid cell coordinate '{code}' for metadata"
                        f" '{metadata.cell_name}'"
                    )
                    continue

                label = self._normalize(cell.value)
                expected_label = self._normalize(metadata.cell_name)
                if label is None:
                    continue
                if expected_label is None or label != expected_label:
                    print(
                        f"Cell {code} in sheet '{sheet.title}' does not"
                        f" match expected name. Expected"
                        f" '{metadata.cell_name}', found '{cell.value}'"
                    )
                    continue

                value_cell = sheet.cell(row=cell.row, column=target_column)
                value = self._stringify(value_cell.value)
                metadata_container.set_value(code, value)
        finally:
            workbook.close()

        return metadata_container

    def write_csv(self, file_path: Path, fieldnames: list[str],
                  rows: list[dict]):
        file_path.parent.mkdir(parents=True, exist_ok=True)
        with file_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

    @staticmethod
    def _normalize(value: object) -> str | None:
        if not isinstance(value, str):
            return None
        cleaned = value.strip()
        return cleaned.casefold() if cleaned else None

    @staticmethod
    def _stringify(value: object) -> str | None:
        if value is None:
            return ""
        text = str(value).strip()
        return text if text else ""
