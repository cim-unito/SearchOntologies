from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from model.metadata_container import MetadataContainer


class MetadataExcelIO:
    """Handle reading and writing metadata values from/to Excel files."""

    def read_metadata_values(
            self, metadata: MetadataContainer, file_path: Path
    ) -> MetadataContainer | None:
        """Populate metadata `values fields using the given Excel file."""
        workbook: Workbook = load_workbook(file_path,
                                           read_only=True,
                                           data_only=True)
        missing_codes = list(set(metadata.codes()))

        try:
            try:
                sheet = workbook[metadata.sheet_name]
            except KeyError:
                print(
                    f"Sheet '{metadata.sheet_name}' not found in"
                    f" '{file_path}'. Available sheets:"
                    f" {', '.join(workbook.sheetnames)}"
                )
                return metadata

            target_column = metadata.column_index
            if target_column < 1:
                print(
                    f"column_index must be a positive integer;"
                    f" got {target_column}"
                )
                return metadata

            for code in missing_codes:
                definition = metadata.get_definition(code)
                if definition is None:
                    continue

                try:
                    cell = sheet[code]
                except ValueError:
                    print(
                        f"Invalid cell coordinate '{code}' for metadata"
                        f" '{definition.cell_name}'"
                    )
                    continue

                label = self._normalize(cell.value)
                expected_label = self._normalize(definition.cell_name)
                if label is None:
                    continue
                if expected_label is None or label != expected_label:
                    print(
                        f"Cell {code} in sheet '{sheet.title}' does not"
                        f" match expected name. Expected"
                        f" '{definition.cell_name}', found '{cell.value}'"
                    )
                    continue

                value_cell = sheet.cell(row=cell.row, column=target_column)
                value = self._stringify(value_cell.value)
                metadata.set_value(code, value)
        finally:
            workbook.close()

        return metadata

    def write_metadata_values(self):
        pass

    def _normalize(self, value: object) -> str | None:
        if not isinstance(value, str):
            return None
        cleaned = value.strip()
        return cleaned.casefold() if cleaned else None

    def _stringify(self, value: object) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text.casefold() if text else None
